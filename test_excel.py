import openpyxl
from typing import List, Optional

def xlsx_to_string(
    filename: str,
    orientation: str = "col",
    col_order: Optional[List[str]] = None,
    doc_per_sheet: bool = False,
    max_rows_per_str: int = 100
) -> List[str]:
    """
    从 Excel 文件中提取文本内容。支持按列或按行聚合数据，并且可以指定表头锚点来定位数据区域。还支持将每个工作表作为独立文档处理。

    主要功能包括：
        1. 定位表头：通过指定 col_order 中的列名，自动识别表头所在行，并以此为基准提取数据。
        2. 数据聚合：根据 orientation 参数选择按列聚合（每列为一个文本块）或按行聚合（每行为一个文本块）。
        3. 分段控制：通过 max_rows_per_str 参数控制每个文本块包含的数据行数，避免单个文本块过大导致处理困难。
        4. 多工作表支持：通过 doc_per_sheet 参数决定是否将每个工作表作为独立文档处理，适用于需要分开分析不同工作表内容的场景。
        5. 支持本地文件和 URL 输入，方便灵活。
        6. 支持合并单元格处理，将父单元格内容合并到子单元格列名中。

    参数:
        - filename (str): Excel 文件的路径或 URL。
        - orientation (str): 数据聚合方式，"col" 表示按列聚合，"row" 表示按行聚合。默认为 "col"。
        - col_order (List[str], optional): 用于定位表头的列名列表，函数会自动寻找包含这些列名的行作为表头。默认为 None，即不使用表头定位。
        - doc_per_sheet (bool): 是否将每个工作表作为独立文档处理。默认为 False。
        - max_rows_per_str (int): 每个文本块包含的最大数据行数，默认为 100。

    返回:
        List[str]: 提取的文本内容列表，每个元素对应一个文本块。
    """
    import openpyxl
    import io, requests

    max_rows_per_str = int(max_rows_per_str)
    # 1. 获取文件流
    file_stream = None
    if filename.startswith(("http://", "https://")):
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        response = requests.get(filename, headers=headers, timeout=30)
        response.raise_for_status()
        file_stream = io.BytesIO(response.content)
    else:
        with open(filename, 'rb') as f:
            file_stream = io.BytesIO(f.read())
    
    # 2. 加载工作簿（注意：不使用 read_only 模式以支持合并单元格读取）
    wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=False)

    def process_sheet(sheet) -> List[str]:
        # 获取合并单元格信息
        merged_ranges = list(sheet.merged_cells.ranges)
        
        # 构建合并单元格映射：(row, col) -> 父单元格值
        merged_cell_map = {}
        for merged_range in merged_ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            parent_value = sheet.cell(min_row, min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_map[(row, col)] = parent_value
        
        print(f"DEBUG: 工作表名称: {sheet.title}")
        print(f"DEBUG: 合并单元格数量: {len(merged_ranges)}")
        
        # 读取所有行数据
        all_rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                        min_col=1, max_col=sheet.max_column))
        
        print(f"DEBUG: 总行数: {len(all_rows)}")
        
        if not all_rows:
            return []
        
        # 1. 寻找表头锚点行
        header_row_idx = None
        if col_order:
            for idx, row in enumerate(all_rows):
                row_str_values = [str(c.value).strip() if c.value is not None else "" for c in row]
                print(f"DEBUG: 第 {idx} 行内容: {row_str_values[:5]}...")  # 只打印前5列
                if any(key in row_str_values for key in col_order):
                    header_row_idx = idx
                    print(f"DEBUG: 找到表头行: {idx}")
                    break
        else:
            header_row_idx = 0
        
        if header_row_idx is None:
            print("DEBUG: 未找到表头行")
            return []
        
        # 2. 检测多行表头
        header_end_idx = header_row_idx
        for check_idx in range(header_row_idx + 1, min(header_row_idx + 5, len(all_rows))):
            check_row = all_rows[check_idx]
            # 检查这一行是否包含表头的一部分（通过检查合并单元格）
            is_header_row = False
            for col_idx in range(len(check_row)):
                actual_row = check_idx + 1
                actual_col = col_idx + 1
                # 如果当前单元格在合并区域内，且合并区域的起始行 <= header_row_idx + 1
                if (actual_row, actual_col) in merged_cell_map:
                    # 查找该合并区域的起始行
                    for mr in merged_ranges:
                        if (actual_row >= mr.min_row and actual_row <= mr.max_row and
                            actual_col >= mr.min_col and actual_col <= mr.max_col):
                            if mr.min_row <= header_row_idx + 1:
                                is_header_row = True
                                break
                if is_header_row:
                    break
            
            if is_header_row:
                header_end_idx = check_idx
                print(f"DEBUG: 检测到多行表头，扩展到第 {check_idx} 行")
            else:
                break
        
        # 3. 构建列名（合并多行表头）
        num_cols = len(all_rows[header_row_idx])
        final_headers = []
        
        for col_idx in range(num_cols):
            header_parts = []
            for row_idx in range(header_row_idx, header_end_idx + 1):
                cell = all_rows[row_idx][col_idx]
                actual_row = row_idx + 1
                actual_col = col_idx + 1
                
                # 获取单元格值（考虑合并单元格）
                if (actual_row, actual_col) in merged_cell_map:
                    cell_value = merged_cell_map[(actual_row, actual_col)]
                else:
                    cell_value = cell.value
                
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    if cell_str and cell_str not in header_parts:
                        header_parts.append(cell_str)
            
            # 合并多行表头
            if header_parts:
                final_headers.append(" ".join(header_parts))
            else:
                final_headers.append(f"Col_{col_idx + 1}")
        
        print(f"DEBUG: 最终列名: {final_headers}")
        
        # 4. 提取数据行
        data_start_idx = header_end_idx + 1
        data_rows = []
        for row in all_rows[data_start_idx:]:
            row_values = [cell.value for cell in row]
            # 跳过完全空行
            if any(v is not None for v in row_values):
                data_rows.append(row_values)
        
        print(f"DEBUG: 数据行数: {len(data_rows)}")
        
        if not data_rows:
            return []

        # 5. 数据分段逻辑
        chunked_results = []
        
        for i in range(0, len(data_rows), max_rows_per_str):
            chunk = data_rows[i : i + max_rows_per_str]
            current_chunk_str = []

            if orientation == "col":
                # 按列聚合
                for col_idx, h_name in enumerate(final_headers):
                    values = [str(r[col_idx]) for r in chunk if col_idx < len(r) and r[col_idx] is not None]
                    if values:
                        current_chunk_str.append(f"{h_name}:\n" + "\n".join(values))
                        current_chunk_str.append("")
            else:
                # 按行聚合
                for r in chunk:
                    row_parts = [f"{final_headers[k]}: {r[k]}" 
                                 for k in range(len(final_headers)) 
                                 if k < len(r) and r[k] is not None]
                    if row_parts:
                        current_chunk_str.append(" | ".join(row_parts))

            result_str = "\n".join(current_chunk_str).strip()
            if result_str:
                chunked_results.append(result_str)

        return chunked_results

    try:
        final_list = []
        if doc_per_sheet:
            for sh in wb.worksheets:
                final_list.extend(process_sheet(sh))
        else:
            final_list.extend(process_sheet(wb.active))
        return final_list
    finally:
        wb.close()

if __name__ == "__main__":
    # Example usage
    result = xlsx_to_string("https://gxw.xianyang.gov.cn/xwzx/tzgg/202312/P020231214614546500860.xlsx", orientation="row", col_order=["企业名称", "序号"], doc_per_sheet=True, max_rows_per_str=50)
    print(f"\n总共生成 {len(result)} 个文本块\n")
    for idx, res in enumerate(result):
        print(f"--- Chunk {idx + 1} ---")
        print(res)
        print("\n")